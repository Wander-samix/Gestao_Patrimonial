# interface/controllers/produto_controller.py

import csv
import json
import xml.etree.ElementTree as ET
from datetime import date, datetime
from itertools import zip_longest

from django.shortcuts            import render, redirect, get_object_or_404
from django.urls                 import reverse
from django.http                 import HttpResponse, JsonResponse
from django.core.files.storage   import FileSystemStorage
from django.contrib              import messages
from django.contrib.auth         import authenticate, login, logout, get_user_model, update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms   import PasswordChangeForm
from django.core.paginator       import Paginator
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import csrf_exempt
from django.db                   import transaction
from django.db.models            import Q, Sum, Max, Min, F, DecimalField
from django.db.utils             import IntegrityError
from django.utils.dateparse      import parse_date
from django.utils.timezone       import now
from django.core.mail            import send_mail
from django.core.serializers.json import DjangoJSONEncoder
from django.conf                 import settings
import requests
from collections import defaultdict

from openpyxl                    import Workbook
from openpyxl.styles             import Font, Alignment, PatternFill
from openpyxl.utils              import get_column_letter

from core.models import (
    Produto, Fornecedor, Area, ItemPedido, Pedido,
    MovimentacaoEstoque, ConfiguracaoEstoque,
    LogAcao, SessionLog, SaidaProdutoPorPedido
)
from interface.forms.forms import (
    AreaForm, ConfiguracaoEstoqueForm,
    ProfileForm, ProdutoForm
)

User = get_user_model()
PENDING_STATUSES = ['aguardando_aprovacao', 'aprovado', 'separado']

def calcular_estoque_disponivel(codigo_barras, area_id, data_limite=None):
    total = Produto.objects.filter(
        codigo_barras=codigo_barras,
        area_id=area_id
    ).aggregate(s=Sum('quantidade'))['s'] or 0

    reservas = ItemPedido.objects.filter(
        produto__codigo_barras=codigo_barras,
        produto__area_id=area_id,
        pedido__status__in=PENDING_STATUSES
    )
    if data_limite:
        reservas = reservas.filter(pedido__data_solicitacao__date__lte=data_limite)
    reservado = reservas.aggregate(s=Sum('quantidade'))['s'] or 0

    return max(total - reservado, 0)


def login_view(request):
    if request.user.is_authenticated:
        return redirect('lista_produtos')
    if request.method == "POST":
        user = authenticate(
            request,
            username=request.POST.get('username'),
            password=request.POST.get('password')
        )
        if user:
            login(request, user)
            return redirect('lista_produtos')
        messages.error(request, "Credenciais inválidas")
    return render(request, 'core/login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
def lista_produtos(request):
    busca         = request.GET.get('busca', '').strip()
    status        = request.GET.get('filtro_status', '')
    filtro_area   = request.GET.get('filtro_area', '')
    ordenar_por   = request.GET.get('ordenar_por', 'id')
    ordem         = request.GET.get('ordem', 'asc')
    somente_baixo = request.GET.get('estoque_baixo') == '1'

    # 1) filtros iniciais
    produtos = Produto.objects.all()
    if busca:
        produtos = produtos.filter(
            Q(descricao__icontains=busca) |
            Q(codigo_barras__icontains=busca)
        )
    if status:
        produtos = produtos.filter(status=status)
    if filtro_area:
        produtos = produtos.filter(area__nome__iexact=filtro_area)
    if ordenar_por in ['id', 'validade', 'descricao', 'criado_em']:
        exp = f"-{ordenar_por}" if ordem == 'desc' else ordenar_por
        produtos = produtos.order_by(exp)

    configs = ConfiguracaoEstoque.objects.all()
    minimo_pct_por_area = {
        cfg.area_id: cfg.estoque_minimo
        for cfg in configs
        if cfg.area_id
    }
    DEFAULT_MINIMO_PCT = 50
    hoje = now().date()

    # 2) reservas pendentes até hoje
    pendentes = (
        ItemPedido.objects
            .filter(
                produto__codigo_barras__in=produtos.values_list('codigo_barras', flat=True),
                produto__area_id__in=produtos.values_list('area_id', flat=True),
                pedido__status__in=PENDING_STATUSES,
                pedido__data_solicitacao__date__lte=hoje
            )
            .values('produto__codigo_barras', 'produto__area_id')
            .annotate(total=Sum('quantidade'))
    )
    reservas = {
        (r['produto__codigo_barras'], r['produto__area_id']): r['total']
        for r in pendentes
    }

    # 3) Calcula totais por grupo (codigo_barras, area_id)
    grupos = defaultdict(lambda: {'real': 0, 'reservado': 0})
    for p in produtos:
        key = (p.codigo_barras, p.area_id)
        grupos[key]['real'] += p.quantidade

    for key, total in reservas.items():
        grupos[key]['reservado'] += total

    # 4) Processa produtos
    produtos_filtrados = []
    for p in produtos.order_by('validade', 'criado_em'):
        key = (p.codigo_barras, p.area_id)
        real = p.quantidade
        reservado = min(real, reservas.get(key, 0))
        reservas[key] = reservas.get(key, 0) - reservado
        disponivel = real - reservado

        # totais do grupo
        total_real_grupo = grupos[key]['real']
        total_reservado_grupo = grupos[key]['reservado']
        disponivel_total_grupo = total_real_grupo - total_reservado_grupo
        disponivel_total_grupo = max(disponivel_total_grupo, 0)

        # Para o template
        p.estoque_real = real
        p.estoque_reservado = reservado
        p.estoque_disponivel_projetado = disponivel
        p.disponivel_total_grupo = disponivel_total_grupo

        # estoque mínimo da área
        pct_minimo = minimo_pct_por_area.get(p.area_id, DEFAULT_MINIMO_PCT)
        threshold = total_real_grupo * (pct_minimo / 100.0)

        p.estoque_baixo = disponivel_total_grupo <= threshold

        # AGORA: percentual do disponível em relação ao total real do grupo
        p.percentual_estoque = (disponivel_total_grupo / total_real_grupo) * 100 if total_real_grupo else 0

        if not somente_baixo or p.estoque_baixo:
            produtos_filtrados.append(p)

    # Conta total de produtos e filtrados para exibir no rodapé se desejar
    total_produtos = produtos.count()
    total_filtrados = len(produtos_filtrados)

    return render(request, 'core/lista_produtos.html', {
        'produtos': produtos_filtrados,
        'busca': busca,
        'filtro_status': status,
        'filtro_area': filtro_area,
        'ordenar_por': ordenar_por,
        'ordem': ordem,
        'areas': Area.objects.all(),
        'fornecedores': Fornecedor.objects.all(),
        'estoque_baixo_aplicado': somente_baixo,
        'total_produtos': total_produtos,
        'total_filtrados': total_filtrados,
    })


@login_required
@require_POST
def upload_nfe(request):
    """
    Recebe um XML de NFe, extrai cabeçalho + itens, e renderiza a página de
    confirmação (core/confirmar_importacao.html) com os campos ocultos de NFe.
    """
    xml_file = request.FILES.get('arquivo')
    if not xml_file:
        messages.error(request, "Nenhum arquivo enviado.")
        return redirect('cadastro_produtos')

    # 1) Salva temporariamente o XML para parsing
    fs   = FileSystemStorage(location='tmp/')
    name = fs.save(xml_file.name, xml_file)
    path = fs.path(name)

    try:
        # 2) Faz parsing do XML
        tree = ET.parse(path)
        root = tree.getroot()

        # 3) Detecta namespace (caso o XML tenha)
        ns = ''
        if root.tag.startswith('{'):
            ns = root.tag.split('}')[0] + '}'

        # --- Extrai campos do cabeçalho ---
        ide = root.find(f".//{ns}ide")
        # Número da NFe
        nfe_numero = ide.find(f"{ns}nNF").text if ide is not None and ide.find(f"{ns}nNF") is not None else ''
        # Data de emissão (formato YYYY-MM-DD)
        data_emissao_raw = ide.find(f"{ns}dEmi").text if ide is not None and ide.find(f"{ns}dEmi") is not None else None
        data_emissao = datetime.strptime(data_emissao_raw, "%Y-%m-%d").date() if data_emissao_raw else None

        # Totais: valor total (vNF) e peso total (vPeso) dentro de ICMSTot
        icmstot = root.find(f".//{ns}ICMSTot")
        valor_total = Decimal('0.00')
        peso = Decimal('0.00')
        if icmstot is not None:
            vnf = icmstot.find(f"{ns}vNF")
            vpr = icmstot.find(f"{ns}vPeso")
            if vnf is not None:
                valor_total = Decimal(vnf.text.replace(',', '.'))
            if vpr is not None:
                peso = Decimal(vpr.text.replace(',', '.'))

        # Emitente (fornecedor) — CNPJ e xNome
        emit = root.find(f".//{ns}emit")
        cnpj_fornecedor = emit.find(f"{ns}CNPJ").text if emit is not None and emit.find(f"{ns}CNPJ") is not None else ''
        nome_fornecedor = emit.find(f"{ns}xNome").text if emit is not None and emit.find(f"{ns}xNome") is not None else ''

        # --- Extrai cada item (det/prod) ---
        produtos_extraidos = []
        dets = root.findall(f".//{ns}det")
        for det in dets:
            prod = det.find(f"{ns}prod")
            if prod is None:
                continue

            codigo_barras   = prod.find(f"{ns}cProd").text if prod.find(f"{ns}cProd") is not None else ''
            descricao       = prod.find(f"{ns}xProd").text if prod.find(f"{ns}xProd") is not None else ''
            qt_text         = prod.find(f"{ns}qCom").text if prod.find(f"{ns}qCom") is not None else '0'
            quantidade      = int(Decimal(qt_text)) if qt_text else 0
            vuc_text        = prod.find(f"{ns}vUnCom").text if prod.find(f"{ns}vUnCom") is not None else '0.00'
            preco_unitario  = float(Decimal(vuc_text.replace(',', '.'))) if vuc_text else 0.0

            # Validade (quase nunca vem no XML; deixamos em branco)
            validade = ''

            # Lote (vai ficar em branco – o Produto.save() calculará)
            lote = ''

            status = "ativo"

            produtos_extraidos.append({
                'codigo_barras':   codigo_barras,
                'descricao':       descricao,
                'fornecedor_nome': nome_fornecedor,
                'quantidade':      quantidade,
                'preco_unitario':  preco_unitario,
                'validade':        validade,
                'lote':            lote,
                'status':          status,
                'area_id':         None,  # usuário selecionará no template
            })

        # Remove o XML temporário
        fs.delete(name)

        if not produtos_extraidos:
            messages.warning(request, "Nenhum item encontrado no XML da NFe.")
            return redirect('cadastro_produtos')

        # 4) Renderiza template de confirmação, enviando cabeçalho + itens
        return render(request, 'core/confirmar_importacao.html', {
            'produtos':         produtos_extraidos,
            'areas':            Area.objects.all(),
            'nfe_numero':       nfe_numero,
            'data_emissao':     data_emissao,       # no formato date
            'cnpj_fornecedor':  cnpj_fornecedor,
            'valor_total':      valor_total,
            'peso':             peso,
        })

    except ET.ParseError:
        fs.delete(name)
        messages.error(request, 'Estrutura do XML inválida. Verifique o arquivo e tente novamente.')
        return redirect('cadastro_produtos')


@login_required
def novo_produto(request):
    """
    Se vier mais de um codigo_barras no POST, faz importação em massa;
    senão usa ProdutoForm para cadastro individual.
    """
    if request.method == "POST":
        cods = request.POST.getlist("codigo_barras")
        # importação em massa
        if len(cods) > 1:
            count = 0
            for cb, desc, fn, aid, val, qt, pu in zip_longest(
                request.POST.getlist("codigo_barras"),
                request.POST.getlist("descricao"),
                request.POST.getlist("fornecedor_nome"),
                request.POST.getlist("area"),
                request.POST.getlist("validade"),
                request.POST.getlist("quantidade"),
                request.POST.getlist("preco_unitario"),
                fillvalue=""
            ):
                validade    = parse_date(val) if val else None
                quantidade  = int(qt or 0)
                preco       = float(pu.replace(',', '.')) if pu else 0.0
                fornecedor  = None
                if fn:
                    fornecedor, _ = Fornecedor.objects.get_or_create(nome=fn)

                Produto.objects.create(
                    codigo_barras   = cb,
                    descricao       = desc,
                    fornecedor      = fornecedor,
                    area_id         = aid or None,
                    validade        = validade,
                    quantidade      = quantidade,
                    preco_unitario  = preco,
                    criado_por      = request.user,
                )
                count += 1
            messages.success(request, f"{count} produto(s) importado(s) com sucesso!")
            return redirect('lista_produtos')

        # cadastro individual via form
        form = ProdutoForm(request.POST)
        if form.is_valid():
            prod = form.save(commit=False)
            prod.criado_por = request.user
            prod.save()
            messages.success(request, "Produto cadastrado com sucesso!")
            return redirect('lista_produtos')
        messages.error(request, "Por favor corrija os erros abaixo.")
    else:
        form = ProdutoForm()

    return render(request, "core/cadastro_produtos.html", {
        "form": form,
        "areas": Area.objects.all(),
        "fornecedores": Fornecedor.objects.all(),
    })


@login_required
def excluir_produto(request, produto_id):
    produto = get_object_or_404(Produto, id=produto_id)
    produto.delete()
    messages.success(request, "Produto excluído com sucesso.")
    return redirect('lista_produtos')


@csrf_exempt
@login_required
def bulk_delete_produtos(request):
    """
    Recebe JSON {"ids":[...]} e exclui produtos.
    """
    if request.method != 'POST':
        return JsonResponse({'sucesso': False, 'erro': 'Método não permitido'}, status=405)
    try:
        dados = json.loads(request.body)
        ids   = dados.get("ids", [])
        if not isinstance(ids, list):
            raise ValueError("ids deve ser uma lista")
        Produto.objects.filter(id__in=ids).delete()
        return JsonResponse({'sucesso': True})
    except Exception as e:
        return JsonResponse({'sucesso': False, 'erro': str(e)}, status=400)


@csrf_exempt
@login_required
def salvar_produto_inline(request):
    """
    Cadastra um produto via JSON e retorna id.
    """
    if request.method == 'POST':
        data = json.loads(request.body)
        fornecedor = Fornecedor.objects.get(id=data['fornecedor'])
        area       = Area.objects.get(id=data['area']) if data.get('area') else None

        iguais = Produto.objects.filter(codigo_barras=data['codigo_barras'])
        lote   = (max(int(p.lote) for p in iguais if p.lote.isdigit()) + 1) if iguais else 1

        produto = Produto.objects.create(
            codigo_barras   = data['codigo_barras'],
            descricao       = data['descricao'],
            fornecedor      = fornecedor,
            area            = area,
            lote            = str(lote),
            validade        = data['validade'],
            quantidade      = int(data['quantidade']),
            preco_unitario  = float(data['preco_unitario'].replace(',', '.')),
            criado_por      = request.user
        )
        return JsonResponse({
            'sucesso': True,
            'id': produto.id,
            'fornecedor_nome': fornecedor.nome,
            'area_nome': area.nome if area else ''
        })
    return JsonResponse({'sucesso': False, 'erro': 'Método inválido'}, status=405)


@login_required
def exportar_produtos_excel(request):
    busca = request.GET.get('busca', '').strip()
    status = request.GET.get('filtro_status', '')
    qs = Produto.objects.select_related('area', 'fornecedor')
    if busca:
        qs = qs.filter(
            Q(descricao__icontains=busca) |
            Q(codigo_barras__icontains=busca)
        )
    if status in ['ativo', 'inativo']:
        qs = qs.filter(status=status)

    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"

    # Cabeçalho com timestamp
    ts = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    ws.merge_cells('A1:K1')
    cell = ws['A1']
    cell.value = f"Exportado em: {ts}"
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='right')

    headers = [
        'Código de Barras','Lote','Descrição','Fornecedor','Área',
        'Validade','Quantidade','Estoque Real','Reservado','Disponível','Preço Unitário'
    ]
    header_fill = PatternFill('solid', fgColor="4F81BD")
    header_font = Font(bold=True, color='FFFFFF')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for idx, title in enumerate(headers, start=1):
        c = ws.cell(row=2, column=idx, value=title)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center

    # -- calcular Reservado e Disponível para cada produto --  
    from interface.controllers.produto_controller import PENDING_STATUSES

    for row_idx, produto in enumerate(qs, start=3):
        reservado = (
            ItemPedido.objects
            .filter(
                produto__codigo_barras=produto.codigo_barras,
                produto__area_id=produto.area_id,
                pedido__status__in=PENDING_STATUSES
            )
            .aggregate(s=Sum('quantidade'))['s'] or 0
        )
        estoque_real = produto.quantidade
        disponivel = max(estoque_real - reservado, 0)

        ws.cell(row=row_idx, column=1, value=produto.codigo_barras)
        ws.cell(row=row_idx, column=2, value=produto.lote)
        ws.cell(row=row_idx, column=3, value=produto.descricao)
        ws.cell(row=row_idx, column=4, value=getattr(produto.fornecedor, 'nome', ''))
        ws.cell(row=row_idx, column=5, value=produto.area.nome if produto.area else "")
        ws.cell(row=row_idx, column=6, value=produto.validade.strftime('%d/%m/%Y') if produto.validade else "")
        ws.cell(row=row_idx, column=7, value=estoque_real)
        ws.cell(row=row_idx, column=8, value=estoque_real)
        ws.cell(row=row_idx, column=9, value=reservado)
        ws.cell(row=row_idx, column=10, value=disponivel)
        ws.cell(row=row_idx, column=11, value=float(produto.preco_unitario))

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename="produtos.xlsx"'
    wb.save(resp)
    return resp


@login_required
@user_passes_test(lambda u: u.papel == 'admin')
def deletar_produto(request, produto_id):
    produto = get_object_or_404(Produto, id=produto_id)
    produto.delete()
    messages.success(request, "Produto excluído com sucesso.")
    return redirect('lista_produtos')


@login_required
def cadastro_produtos(request):
    """
    Recebe POST de core/confirmar_importacao.html e cadastra produtos em massa.
    Também trata fallback individual.
    Ao final, redireciona para lista de produtos.
    """
    areas = Area.objects.all()
    fornecedores = Fornecedor.objects.all()

    if request.method == "POST":
        try:
            # --- 1) Importação em massa via NFe ---
            codigos       = request.POST.getlist("codigo_barras")
            descricoes    = request.POST.getlist("descricao")
            nomes_forn    = request.POST.getlist("fornecedor_nome")
            area_ids      = request.POST.getlist("area_id")
            lotes         = request.POST.getlist("lote")
            validades     = request.POST.getlist("validade")
            quantidades   = request.POST.getlist("quantidade")
            precos        = request.POST.getlist("preco_unitario")
            statuses      = request.POST.getlist("status")
            # Listagem de NFe (a cada linha terá o mesmo número, mas pegamos via getlist)
            nfe_nums      = request.POST.getlist("nfe_numero")

            # Se houver ao menos um código (importação em massa)
            if any(codigos):
                created = 0
                for (
                    cb, desc, fn, aid, lote, val, qt, pr, st, nfe
                ) in zip_longest(
                    codigos,
                    descricoes,
                    nomes_forn,
                    area_ids,
                    lotes,
                    validades,
                    quantidades,
                    precos,
                    statuses,
                    nfe_nums,
                    fillvalue=""
                ):
                    # --- 1.1) Parse dos campos ---
                    # Validade → date
                    validade_date = None
                    if val:
                        try:
                            validade_date = datetime.strptime(val, "%Y-%m-%d").date()
                        except:
                            try:
                                validade_date = datetime.strptime(val, "%d/%m/%Y").date()
                            except:
                                validade_date = None

                    # Quantidade → int
                    try:
                        quantidade = int(qt)
                    except:
                        quantidade = 0

                    # Preço Unitário → float (substitui vírgula por ponto)
                    preco_unitario = 0.0
                    if pr:
                        preco_unitario = float(pr.replace(',', '.'))

                    # Status (ou “ativo” se vazio)
                    status_final = st or "ativo"

                    # --- 1.2) Cria ou recupera Fornecedor ---
                    if fn:
                        fornecedor, _ = Fornecedor.objects.get_or_create(nome=fn)
                    else:
                        fornecedor = None

                    # --- 1.3) Cria Produto (lote automático no save()) ---
                    produto = Produto(
                        nfe_numero      = nfe,       # sempre o mesmo número
                        codigo_barras   = cb,
                        descricao       = desc,
                        fornecedor      = fornecedor,
                        area_id         = aid or None,
                        lote            = None,      # None → save() calcula
                        validade        = validade_date,
                        quantidade      = quantidade,
                        preco_unitario  = preco_unitario,
                        status          = status_final,
                        criado_por      = request.user,
                    )
                    produto.save()
                    created += 1

                messages.success(request, f"{created} produto(s) importado(s) com sucesso!")
                return redirect("lista_produtos")

            # --- 2) Fallback de cadastro individual via formulário HTML ---
            codigo_barras = request.POST.get("codigo_barras")
            if codigo_barras:
                descricao      = request.POST.get("descricao", "")
                fornecedor_id  = request.POST.get("fornecedor")
                area_id        = request.POST.get("area_id")
                lote           = request.POST.get("lote", "")
                validade_str   = request.POST.get("validade")
                quantidade     = int(request.POST.get("quantidade", 0))
                preco_unitario = float(request.POST.get("preco_unitario", "0").replace(',', '.'))
                status_final   = request.POST.get("status") or "ativo"

                validade_date = parse_date(validade_str) if validade_str else None

                if not fornecedor_id or not area_id:
                    messages.error(request, "Fornecedor e área são obrigatórios.")
                else:
                    Produto.objects.create(
                        codigo_barras   = codigo_barras,
                        descricao       = descricao,
                        fornecedor_id   = fornecedor_id,
                        area_id         = area_id,
                        lote            = lote,
                        validade        = validade_date,
                        quantidade      = quantidade,
                        preco_unitario  = preco_unitario,
                        status          = status_final,
                        criado_por      = request.user,
                    )
                    messages.success(request, "Produto cadastrado com sucesso!")
                return redirect("lista_produtos")

        except IntegrityError:
            messages.error(request, "Erro ao cadastrar produto. Verifique os dados.")
            return redirect("cadastro_produtos")

    # GET: exibe o formulário de cadastro/importação
    return render(request, "core/cadastro_produtos.html", {
        "areas":        areas,
        "fornecedores": fornecedores,
    })


@login_required
def novo_produto_individual(request):
    if request.method == 'POST':
        form = ProdutoForm(request.POST)
        if form.is_valid():
            prod = form.save(commit=False)
            prod.criado_por = request.user
            prod.save()
            messages.success(request, "Produto cadastrado com sucesso!")
            return redirect('lista_produtos')
    else:
        form = ProdutoForm()

    api_template = reverse('buscar_nome_produto', kwargs={'codigo': 'DUMMY'})
    api_base     = api_template.replace('DUMMY/', '')

    return render(request, 'core/produto_form.html', {
        'form': form,
        'titulo': 'Novo Produto',
        'produto': None,
        'areas': Area.objects.all(),
        'fornecedores': Fornecedor.objects.all(),
        'api_base': api_base,
    })


@login_required
def editar_produto(request, id):
    produto = get_object_or_404(Produto, pk=id)
    if request.method == 'POST':
        form = ProdutoForm(request.POST, instance=produto)
        if form.is_valid():
            form.save()
            messages.success(request, "Produto atualizado com sucesso.")
            return redirect('lista_produtos')
    else:
        form = ProdutoForm(instance=produto)

    api_template = reverse('buscar_nome_produto', kwargs={'codigo': 'DUMMY'})
    api_base     = api_template.replace('DUMMY', '')

    return render(request, 'core/produto_form.html', {
        'form': form,
        'titulo': 'Editar Produto',
        'produto': produto,
        'areas': Area.objects.all(),
        'fornecedores': Fornecedor.objects.all(),
        'api_base': api_base,
    })


def buscar_nome_produto_por_codigo(codigo_barras):
    """
    Lê a COSMOS_API_KEY do settings e busca /gtins/{codigo}.json.
    """
    url = f"https://api.cosmos.bluesoft.com.br/gtins/{codigo_barras}.json"
    headers = {
        "X-Cosmos-Token": settings.COSMOS_API_KEY,
        "Content-Type": "application/json",
        "User-Agent": "Cosmos-API-Request",
    }
    try:
        resp = requests.get(url, headers=headers, timeout=5)
        if resp.status_code == 200:
            return resp.json().get("description", "Descrição não encontrada")
        if resp.status_code == 404:
            return "Produto não encontrado"
        return f"Erro: {resp.status_code}"
    except Exception:
        return "Erro na consulta"


@login_required
@require_GET
def buscar_nome_produto_view(request, codigo):
    nome = buscar_nome_produto_por_codigo(codigo)
    return JsonResponse({
        "codigo": codigo,
        "nome_produto": nome
    })

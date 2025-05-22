from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.utils import IntegrityError
from django.utils.dateparse import parse_date
from django.utils.timezone import now
from django.core.files.storage import FileSystemStorage
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.db.models.functions import TruncMonth
from django.core.mail import send_mail
from django.core.serializers.json import DjangoJSONEncoder
import xml.etree.ElementTree as ET
import json
import requests
from openpyxl import Workbook
from django.db.models import Q, Sum, Max, Min, F, DecimalField
from django.views.decorators.http import require_GET
from itertools import zip_longest
from core.models import (
    Produto, Fornecedor, Usuario, MovimentacaoEstoque,
    Area, Pedido, ItemPedido, LogAcao,
    ConfiguracaoEstoque)
from ..forms.forms import AreaForm, ConfiguracaoEstoqueForm
from django.db import transaction
from django import template
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from ..forms.forms import ProfileForm
from ..forms.forms import ProdutoForm
import csv
from datetime import date, datetime
from django.shortcuts import render
from core.models import SessionLog
from django.core.paginator import Paginator
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.db.models import OuterRef, Subquery
from core.models import SaidaProdutoPorPedido


register = template.Library()
User = get_user_model()

PENDING_STATUSES = [
    'aguardando_aprovacao',
    'aprovado',
    'separado',
]


def lista_sessoes(request):
    sessoes = SessionLog.objects.all()
    return render(request, 'core/lista_sessoes.html', {'sessoes': sessoes})

@login_required
@user_passes_test(lambda u: u.is_superuser)

def exportar_sessoes_excel(request):
    # busca todas as sessões
    sessoes = SessionLog.objects.select_related('user').order_by('-login_time')

    wb = Workbook()
    ws = wb.active
    ws.title = "Sessões de Usuário"

    # cabeçalho
    ws.append(["Usuário", "Login", "Logout", "Duração", "IP"])

    # linhas de dados
    for s in sessoes:
        ws.append([
            s.user.username,
            s.login_time.strftime("%d/%m/%Y %H:%M:%S"),
            s.logout_time.strftime("%d/%m/%Y %H:%M:%S") if s.logout_time else "—",
            str(s.duration) if s.duration else "—",
            s.ip or "—",
        ])

    # monta resposta
    filename = f"sessoes_{date.today().isoformat()}.xlsx"
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


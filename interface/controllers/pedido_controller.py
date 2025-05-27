# interface/controllers/pedido_controller.py

from datetime import datetime
from io import BytesIO
from itertools import zip_longest

from django.shortcuts                import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib                  import messages
from django.db                       import transaction
from django.db.models                import Sum, Max
from django.http                     import HttpResponse
from django.utils.timezone           import now
from openpyxl                        import Workbook
from openpyxl.styles                 import Font, PatternFill, Alignment
from openpyxl.utils                  import get_column_letter

from core.models    import Pedido, ItemPedido, Produto, Area

PENDING_STATUSES = ['aguardando_aprovacao', 'aprovado', 'separado']


def is_admin_tecnico(user):
    return user.is_authenticated and user.papel in ('admin', 'tecnico')


def calcular_estoque_disponivel(codigo_barras, area_id, data_limite=None):
    """
    Retorna o estoque disponível (quantidade total - pendentes)
    até data_limite, num combo código+área.
    """
    total = Produto.objects.filter(
        codigo_barras=codigo_barras, area_id=area_id
    ).aggregate(s=Sum('quantidade'))['s'] or 0

    reservas = ItemPedido.objects.filter(
        produto__codigo_barras=codigo_barras,
        produto__area_id=area_id,
        pedido__status__in=PENDING_STATUSES
    )
    if data_limite:
        reservas = reservas.filter(pedido__data_solicitacao__lte=data_limite)

    reservado = reservas.aggregate(s=Sum('quantidade'))['s'] or 0
    return max(total - reservado, 0)


@login_required
def lista_pedidos(request):
    status      = request.GET.get("status", "")
    ordenar_por = request.GET.get("ordenar_por", "-data_solicitacao")
    page        = request.GET.get("page")

    qs = Pedido.objects.all() if is_admin_tecnico(request.user) \
         else Pedido.objects.filter(usuario=request.user)

    if status:
        qs = qs.filter(status=status)

    qs = qs.order_by(ordenar_por)
    from django.core.paginator import Paginator
    pedidos = Paginator(qs, 10).get_page(page)

    return render(request, 'core/lista_pedidos.html', {
        'pedidos': pedidos,
        'status_selecionado': status,
        'ordenar_por': ordenar_por.lstrip('-'),
        'ordem': 'desc' if ordenar_por.startswith('-') else 'asc',
    })


@login_required
@transaction.atomic
def novo_pedido(request):
    if request.method == "POST":
        # 1) Gera código PD00001, PD00002...
        ultimo = Pedido.objects.aggregate(max_id=Max("id"))["max_id"] or 0
        codigo = f"PD{str(ultimo + 1).zfill(5)}"

        # 2) Define status inicial e quem aprovou
        if request.user.papel in ("admin", "tecnico"):
            status_inicial = "aprovado"
            aprovado_por   = request.user
            data_aprovacao = now()
        else:
            status_inicial = "aguardando_aprovacao"
            aprovado_por   = None
            data_aprovacao = None

        # 3) Cria o pedido
        data_necessaria = request.POST.get('data_necessaria') or None
        pedido = Pedido.objects.create(
            codigo=codigo,
            usuario=request.user,
            status=status_inicial,
            aprovado_por=aprovado_por,
            data_aprovacao=data_aprovacao,
            data_necessaria=data_necessaria,
        )

        # 4) Loop pelos itens vindos do form
        hoje        = now().date()
        area_ids    = request.POST.getlist('area_id')
        prod_ids    = request.POST.getlist('produto_id')
        quant_list  = request.POST.getlist('quantidade')
        obs_list    = request.POST.getlist('observacao')

        for area_id, prod_id, qt, obs in zip_longest(
                area_ids, prod_ids, quant_list, obs_list,
                fillvalue=None):

            if not (area_id and prod_id and qt):
                continue

            quantidade = int(qt)
            # trava o registro pra consistência
            prod = get_object_or_404(
                Produto.objects.select_for_update(), pk=int(prod_id)
            )
            if prod.area_id != int(area_id):
                messages.error(request, f"Área inconsistente para {prod.descricao}.")
                raise ValueError("Área inconsistente")

            disponível = calcular_estoque_disponivel(
                prod.codigo_barras, prod.area_id, data_limite=hoje
            )
            if disponível < quantidade:
                messages.error(
                    request,
                    f"Estoque insuficiente para {prod.descricao}: "
                    f"{disponível} disponível, {quantidade} solicitado."
                )
                raise ValueError("Estoque insuficiente")

            ItemPedido.objects.create(
                pedido=pedido,
                produto=prod,
                quantidade=quantidade,
                observacao=(obs or "").strip(),
                estoque_no_pedido=disponível
            )

        messages.success(request, "Pedido criado com sucesso!")
        return redirect('lista_pedidos')

    # GET → apenas fornece lista de áreas (o JS buscará produtos via AJAX)
    áreas = Area.objects.all() if is_admin_tecnico(request.user) \
           else request.user.areas.all()

    return render(request, 'core/novo_pedido.html', {
        'areas': áreas,
    })


@login_required
def detalhes_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    return render(request, 'core/detalhes_pedido.html', {
        'pedido': pedido,
        'eh_admin_tecnico': is_admin_tecnico(request.user),
    })


@login_required
def editar_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    if request.method == "POST":
        novo_status     = request.POST.get('status')
        data_necessaria = request.POST.get('data_necessaria')
        if novo_status:
            pedido.status = novo_status
        if data_necessaria:
            try:
                pedido.data_necessaria = datetime.strptime(
                    data_necessaria, '%Y-%m-%d'
                ).date()
            except ValueError:
                pass
        pedido.save()
        messages.success(request, "Pedido atualizado com sucesso.")
        return redirect('lista_pedidos')

    return render(request, 'core/editar_pedido.html', {
        'pedido': pedido,
        'STATUS_CHOICES': Pedido.STATUS_CHOICES,
    })


@login_required
def deletar_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    pedido.delete()
    messages.success(request, "Pedido excluído com sucesso.")
    return redirect('lista_pedidos')


@login_required
def exportar_pedidos_excel(request):
    status = request.GET.get('status')
    qs     = Pedido.objects.select_related('usuario', 'aprovado_por')\
                            .prefetch_related('itens__produto__area')\
                            .order_by('-data_solicitacao')
    if status:
        qs = qs.filter(status=status)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos"

    # Cabeçalho
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="4F81BD")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cols = [
        "Código Pedido", "Data Solicitação", "Solicitante", "Email",
        "Necessário em", "Status", "Aprovado por", "Data Aprovação",
        "Produto", "Código de Barras", "Área",
        "Qtd Solicitada", "Qtd Liberada", "Estoque no Pedido", "Observação",
    ]
    for i, title in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i, value=title)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align

    row = 2
    for pedido in qs:
        for item in pedido.itens.all():
            prod = item.produto
            ws.cell(row=row, column=1,  value=pedido.codigo)
            ws.cell(row=row, column=2,  value=pedido.data_solicitacao.strftime("%d/%m/%Y %H:%M") if pedido.data_solicitacao else "")
            ws.cell(row=row, column=3,  value=pedido.usuario.username if pedido.usuario else "")
            ws.cell(row=row, column=4,  value=pedido.usuario.email    if pedido.usuario else "")
            ws.cell(row=row, column=5,  value=pedido.data_necessaria.strftime("%d/%m/%Y") if pedido.data_necessaria else "")
            ws.cell(row=row, column=6,  value=pedido.get_status_display())
            ws.cell(row=row, column=7,  value=pedido.aprovado_por.username if pedido.aprovado_por else "")
            ws.cell(row=row, column=8,  value=pedido.data_aprovacao.strftime("%d/%m/%Y %H:%M") if pedido.data_aprovacao else "")

            ws.cell(row=row, column=9,  value=prod.descricao if prod else "")
            ws.cell(row=row, column=10, value=prod.codigo_barras if prod else "")
            ws.cell(row=row, column=11, value=prod.area.nome     if prod and prod.area else "")

            ws.cell(row=row, column=12, value=item.quantidade)
            ws.cell(row=row, column=13, value=item.liberado or item.quantidade)
            ws.cell(row=row, column=14, value=item.estoque_no_pedido or "")
            ws.cell(row=row, column=15, value=item.observacao or "")
            row += 1

    # Ajusta largura das colunas
    for col_cells in ws.columns:
        width = max(len(str(c.value or "")) for c in col_cells) + 2
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width

    # Retorna o arquivo
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp['Content-Disposition'] = 'attachment; filename=pedidos.xlsx'
    return resp

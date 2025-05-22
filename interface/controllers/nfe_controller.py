from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.utils import IntegrityError
from django.utils.dateparse import parse_date
from django.utils.timezone import now
from django.core.files.storage import FileSystemStorage
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.db.models.functions import TruncMonth
from django.core.mail import send_mail
from django.core.serializers.json import DjangoJSONEncoder
import xml.etree.ElementTree as ET
import json
import requests
from openpyxl import Workbook
from django.db.models import Q, Sum, Max, Min, F, DecimalField
from django.views.decorators.http import require_GET
from itertools import zip_longest
from core.models import (
    Produto, Fornecedor, Usuario, MovimentacaoEstoque,
    Area, Pedido, ItemPedido, LogAcao,
    ConfiguracaoEstoque)
from ..forms.forms import AreaForm, ConfiguracaoEstoqueForm
from django.db import transaction
from django import template
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from ..forms.forms import ProfileForm
from ..forms.forms import ProdutoForm
import csv
from datetime import date, datetime
from django.shortcuts import render
from core.models import SessionLog
from django.core.paginator import Paginator
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.db.models import OuterRef, Subquery
from core.models import SaidaProdutoPorPedido


register = template.Library()
User = get_user_model()

PENDING_STATUSES = [
    'aguardando_aprovacao',
    'aprovado',
    'separado',
]


def upload_nfe(request):
    if request.method == "POST" and request.FILES.get("arquivo"):
        xml_file = request.FILES["arquivo"]
        fs = FileSystemStorage()
        filepath = fs.path(fs.save(xml_file.name, xml_file))
        tree = ET.parse(filepath)
        root = tree.getroot()
        ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}

        # 1) EXTRAIR O NÚMERO DA NFe
        nfe_numero = root.findtext(".//nfe:ide/nfe:nNF", default="", namespaces=ns)

        # 2) DADOS DO EMITENTE
        emitente = root.find(".//nfe:emit", ns)
        fornecedor_nome = emitente.findtext("nfe:xNome", default="", namespaces=ns)
        fornecedor_cnpj  = emitente.findtext("nfe:CNPJ", default="", namespaces=ns)
        fornecedor, _ = Fornecedor.objects.get_or_create(
            cnpj=fornecedor_cnpj,
            defaults={"nome": fornecedor_nome, "endereco":"", "telefone":"", "email":""}
        )

        # 3) ITENS
        produtos_extraidos = []
        for det in root.findall(".//nfe:det", ns):
            prod = det.find("nfe:prod", ns)
            if prod is None:
                continue
            produtos_extraidos.append({
                "nfe_numero":      nfe_numero,
                "codigo_barras":   prod.findtext("nfe:cProd",  default="", namespaces=ns),
                "descricao":       prod.findtext("nfe:xProd",  default="", namespaces=ns),
                "quantidade":      prod.findtext("nfe:qCom",   default="0",  namespaces=ns),
                "preco_unitario":  prod.findtext("nfe:vUnCom", default="0.00", namespaces=ns),
                "lote":            "",
                "validade":        "",
                "status":          "ativo",
                "fornecedor_nome": fornecedor.nome,     # passa o nome, não o ID
                "area_id":         None,                # será selecionada no template
            })

        return render(request, "core/confirmar_importacao.html", {
            "produtos":    produtos_extraidos,
            "areas":       Area.objects.all(),
            "nfe_numero":  nfe_numero,
        })

    return render(request, "core/upload_nfe.html")


@login_required

def exportar_produtos_excel(request):
    busca = request.GET.get("busca")
    status = request.GET.get("filtro_status")
    produtos = Produto.objects.all()

    if busca:
        produtos = produtos.filter(
            Q(descricao__icontains=busca) |
            Q(codigo_barras__icontains=busca)
        )
    if status:
        produtos = produtos.filter(status=status)

    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"

    # 1) Carimbo de data/hora na primeira linha
    timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    cell_stamp = ws.cell(row=1, column=1, value=f"Exportado em: {timestamp}")
    cell_stamp.font = Font(bold=True)
    cell_stamp.alignment = Alignment(horizontal="right")

    # 2) Cabeçalho (linha 2)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "Código de Barras", "Lote", "Descrição", "Fornecedor", "Área",
        "Validade", "Quantidade", "Estoque Real", "Reservado", "Disponível",
        "Preço Unitário", "Usuário", "Data de Cadastro"
    ]
    for col_num, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # 3) Dados dos produtos (a partir da linha 3)
    for idx, produto in enumerate(produtos, start=3):
        info = produto.estoque_info(data_limite=None)  # ou passe data se desejar histórico
        ws.cell(row=idx, column=1, value=produto.codigo_barras)
        ws.cell(row=idx, column=2, value=produto.lote)
        ws.cell(row=idx, column=3, value=produto.descricao)
        ws.cell(row=idx, column=4, value=getattr(produto.fornecedor, 'nome', ''))
        ws.cell(row=idx, column=5, value=produto.area.nome if produto.area else "")
        ws.cell(row=idx, column=6, value=produto.validade.strftime('%d/%m/%Y') if produto.validade else "")
        ws.cell(row=idx, column=7, value=produto.quantidade)
        ws.cell(row=idx, column=8, value=info['real'])
        ws.cell(row=idx, column=9, value=info['reservado'])
        ws.cell(row=idx, column=10, value=info['disponivel'])
        ws.cell(row=idx, column=11, value=float(produto.preco_unitario))
        ws.cell(row=idx, column=12, value=produto.criado_por.username if produto.criado_por else "")
        ws.cell(row=idx, column=13, value=produto.criado_em.strftime('%d/%m/%Y %H:%M') if produto.criado_em else "")

    # 4) Ajustar largura das colunas
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # 5) Resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=produtos.xlsx'
    wb.save(response)
    return response

@login_required


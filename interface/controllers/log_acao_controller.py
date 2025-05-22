from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.utils import IntegrityError
from django.utils.dateparse import parse_date
from django.utils.timezone import now
from django.core.files.storage import FileSystemStorage
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.db.models.functions import TruncMonth
from django.core.mail import send_mail
from django.core.serializers.json import DjangoJSONEncoder
import xml.etree.ElementTree as ET
import json
import requests
from openpyxl import Workbook
from django.db.models import Q, Sum, Max, Min, F, DecimalField
from django.views.decorators.http import require_GET
from itertools import zip_longest
from core.models import (
    Produto, Fornecedor, Usuario, MovimentacaoEstoque,
    Area, Pedido, ItemPedido, LogAcao,
    ConfiguracaoEstoque)
from ..forms.forms import AreaForm, ConfiguracaoEstoqueForm
from django.db import transaction
from django import template
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from ..forms.forms import ProfileForm
from ..forms.forms import ProdutoForm
import csv
from datetime import date, datetime
from django.shortcuts import render
from core.models import SessionLog
from django.core.paginator import Paginator
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.db.models import OuterRef, Subquery
from core.models import SaidaProdutoPorPedido


register = template.Library()
User = get_user_model()

PENDING_STATUSES = [
    'aguardando_aprovacao',
    'aprovado',
    'separado',
]


def lista_logs(request):
    if not request.user.is_superuser:
        return redirect('lista_produtos')
    logs = LogAcao.objects.all().order_by('-data_hora')
    return render(request, "core/logs.html", {"logs": logs})

@login_required
@transaction.atomic

def exportar_log_excel(request):
    # opcional: controle de permissão
    if not request.user.is_superuser:
        return redirect('lista_produtos')

    logs = LogAcao.objects.select_related('usuario').order_by('-data_hora')
    wb = Workbook()
    ws = wb.active
    ws.title = "Logs de Auditoria"
    ws.append(["Usuário", "Ação", "Detalhes", "Data/Hora", "IP"])

    for log in logs:
        ws.append([
            log.usuario.username if log.usuario else "—",
            log.acao,
            log.detalhes or "—",
            log.data_hora.strftime("%d/%m/%Y %H:%M:%S"),
            log.ip or "—",
        ])

    filename = f"auditoria_logs_{datetime.date.today().isoformat()}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required

def exportar_logs_excel(request):
    # só superuser ou admin pode baixar
    if not request.user.is_superuser:
        return redirect('lista_produtos')

    logs = LogAcao.objects.select_related('usuario').order_by('-data_hora')

    wb = Workbook()
    ws = wb.active
    ws.title = "Logs de Auditoria"

    # cabeçalho
    ws.append(["Usuário", "Ação", "Detalhes", "Data/Hora", "IP"])

    # linhas
    for log in logs:
        ws.append([
            log.usuario.username if log.usuario else "—",
            log.acao,
            log.detalhes or "—",
            log.data_hora.strftime("%d/%m/%Y %H:%M:%S"),
            log.ip or "—",
        ])

    # gera nome de arquivo com a data de hoje
    filename = f"auditoria_logs_{date.today().isoformat()}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@csrf_exempt
@login_required


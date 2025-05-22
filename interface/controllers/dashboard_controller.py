from datetime import datetime, date
from io import BytesIO

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils.dateparse import parse_date
from django.utils.timezone import now
from django.http import HttpResponse
from django.db.models import Sum, F, DecimalField

from openpyxl import Workbook

from core.models import (
    Produto,
    Pedido,
    Usuario,
    MovimentacaoEstoque,
    LogAcao,
)

PENDING_STATUSES = [
    'aguardando_aprovacao',
    'aprovado',
    'separado',
]

@login_required
def dashboard(request):
    # 1) Parâmetro 'data' (YYYY-MM-DD); usa hoje se ausente ou inválido
    data_str = request.GET.get('data')
    data_mov = parse_date(data_str) or now().date()

    # 2) Contagens básicas
    produtos_count = Produto.objects.count()
    pedidos_count  = Pedido.objects.count()
    usuarios_count = Usuario.objects.count()

    # 3) Pedidos por status
    pedidos_por_status = {
        status: Pedido.objects.filter(status=status).count()
        for status, _ in Pedido.STATUS_CHOICES
    }

    # 4) Entradas totais (quantidade inicial cadastrada)
    entradas_totais = Produto.objects.aggregate(
        total=Sum('quantidade_inicial')
    )['total'] or 0

    # 5) Em estoque (estoque real): soma de estoque_info.real por combo código+área
    estoque_real_total = 0
    combos = Produto.objects.values('codigo_barras', 'area_id').distinct()
    for combo in combos:
        p = Produto.objects.filter(
            codigo_barras=combo['codigo_barras'],
            area_id=combo['area_id']
        ).first()
        if p:
            estoque_real_total += p.estoque_info(data_limite=data_mov)['real']

    # 6) Valor total em estoque (quantidade * preço_unitário)
    valor_total = Produto.objects.aggregate(
        valor=Sum(
            F('quantidade') * F('preco_unitario'),
            output_field=DecimalField(max_digits=20, decimal_places=2)
        )
    )['valor'] or 0

    # 7) Movimentações e logs do dia
    movimentacoes_dia = (
        MovimentacaoEstoque.objects
            .select_related('produto', 'usuario')
            .filter(data__date=data_mov)
            .order_by('-data')
    )
    logs_dia = LogAcao.objects \
        .filter(data_hora__date=data_mov) \
        .order_by('-data_hora')

    return render(request, 'core/dashboard.html', {
        'produtos_count':      produtos_count,
        'pedidos_count':       pedidos_count,
        'usuarios_count':      usuarios_count,
        'pedidos_por_status':  pedidos_por_status,
        'entradas_totais':     entradas_totais,
        'estoque_real_total':  estoque_real_total,
        'valor_total':         valor_total,
        'movimentacoes_dia':   movimentacoes_dia,
        'logs_dia':            logs_dia,
        'data_mov':            data_mov,
        'STATUS_CHOICES':      Pedido.STATUS_CHOICES,
    })


@login_required
def exportar_dashboard_excel(request):
    """
    Gera um arquivo .xlsx com três abas:
      - Resumo (contagens, somas e status)
      - Movimentações do dia
      - Logs do dia
    """
    # reusar mesmo cálculo de data do dashboard
    data_str = request.GET.get('data')
    data_mov = parse_date(data_str) or now().date()

    # recalc indicadores
    produtos_count = Produto.objects.count()
    pedidos_count  = Pedido.objects.count()
    usuarios_count = Usuario.objects.count()
    pedidos_por_status = {
        status: Pedido.objects.filter(status=status).count()
        for status, _ in Pedido.STATUS_CHOICES
    }
    entradas_totais = Produto.objects.aggregate(
        total=Sum('quantidade_inicial')
    )['total'] or 0

    estoque_real_total = 0
    combos = Produto.objects.values('codigo_barras', 'area_id').distinct()
    for combo in combos:
        p = Produto.objects.filter(
            codigo_barras=combo['codigo_barras'],
            area_id=combo['area_id']
        ).first()
        if p:
            estoque_real_total += p.estoque_info(data_limite=data_mov)['real']

    valor_total = Produto.objects.aggregate(
        valor=Sum(
            F('quantidade') * F('preco_unitario'),
            output_field=DecimalField(max_digits=20, decimal_places=2)
        )
    )['valor'] or 0

    movimentacoes = (
        MovimentacaoEstoque.objects
            .select_related('produto', 'usuario')
            .filter(data__date=data_mov)
            .order_by('-data')
    )
    logs = LogAcao.objects \
        .filter(data_hora__date=data_mov) \
        .order_by('-data_hora')

    # montar workbook
    wb = Workbook()

    # aba Resumo
    ws = wb.active
    ws.title = "Resumo"
    ws.append(["Métrica", "Valor"])
    ws.append(["Produtos cadastrados", produtos_count])
    ws.append(["Pedidos cadastrados", pedidos_count])
    ws.append(["Usuários cadastrados", usuarios_count])
    for status, _ in Pedido.STATUS_CHOICES:
        ws.append([f"Pedidos '{status}'", pedidos_por_status[status]])
    ws.append(["Entradas totais", entradas_totais])
    ws.append(["Estoque real total", estoque_real_total])
    ws.append(["Valor total em estoque", float(valor_total)])

    # aba Movimentações
    ws2 = wb.create_sheet("Movimentações")
    ws2.append(["Tipo", "Produto", "Quantidade", "Usuário", "Data/Hora"])
    for m in movimentacoes:
        ws2.append([
            m.get_tipo_display(),
            m.produto.descricao,
            m.quantidade,
            m.usuario.username,
            m.data.strftime("%d/%m/%Y %H:%M"),
        ])

    # aba Logs
    ws3 = wb.create_sheet("Logs")
    ws3.append(["Ação", "Usuário", "Data/Hora"])
    for l in logs:
        ws3.append([
            getattr(l, "acao", ""),
            getattr(l.usuario, "username", ""),
            l.data_hora.strftime("%d/%m/%Y %H:%M"),
        ])

    # preparar resposta
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"dashboard_{data_mov.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

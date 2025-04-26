from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.utils import IntegrityError
from django.utils.dateparse import parse_date
from django.utils.timezone import now
from django.core.files.storage import FileSystemStorage
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.db.models.functions import TruncMonth
from django.core.mail import send_mail
from django.core.serializers.json import DjangoJSONEncoder
import xml.etree.ElementTree as ET
import json
import requests
from openpyxl import Workbook
from django.db.models import Q, Sum, Max, Min, F, DecimalField
from django.views.decorators.http import require_GET
from itertools import zip_longest
from .models import (
    Produto, Fornecedor, Usuario, MovimentacaoEstoque,
    Area, Pedido, ItemPedido, LogAcao,
    ConfiguracaoEstoque)
from .forms import AreaForm, ConfiguracaoEstoqueForm
from django.db import transaction
from django import template
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from .forms import ProfileForm
from .forms import ProdutoForm
import csv
from datetime import date, datetime
from django.shortcuts import render
from .models import SessionLog
from django.core.paginator import Paginator
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.db.models import OuterRef, Subquery
from .models import SaidaProdutoPorPedido
from django.urls import reverse

register = template.Library()
User = get_user_model()

PENDING_STATUSES = [
    'aguardando_aprovacao',
    'aprovado',
    'separado',
]

def calcular_estoque_disponivel(codigo_barras, area_id, data_limite=None):
    total = Produto.objects.filter(
        codigo_barras=codigo_barras,
        area_id=area_id
    ).aggregate(s=Sum('quantidade'))['s'] or 0

    reservas_qs = ItemPedido.objects.filter(
        produto__codigo_barras=codigo_barras,
        produto__area_id=area_id,
        pedido__status__in=PENDING_STATUSES
    )

    if data_limite:
        reservas_qs = reservas_qs.filter(pedido__data_solicitacao__lte=data_limite)

    reservado = reservas_qs.aggregate(s=Sum('quantidade'))['s'] or 0

    return max(total - reservado, 0)


# ---------- Autenticação ----------

def login_view(request):
    if request.user.is_authenticated:
        return redirect('lista_produtos')
    if request.method == "POST":
        user = authenticate(request,
                            username=request.POST.get('username'),
                            password=request.POST.get('password'))
        if user:
            login(request, user)
            return redirect('lista_produtos')
        return render(request, 'core/login.html', {'error': 'Credenciais inválidas'})
    return render(request, 'core/login.html')

def logout_view(request):
    logout(request)
    return redirect('login')


# core/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q, Sum
from django.utils.timezone import now
from .models import Produto, Area, Fornecedor, ItemPedido, ConfiguracaoEstoque, PENDING_STATUSES

@login_required
def lista_produtos(request):
    busca         = request.GET.get('busca', '')
    status        = request.GET.get('filtro_status', '')
    filtro_area   = request.GET.get('filtro_area', '')
    ordenar_por   = request.GET.get('ordenar_por', 'id')
    ordem         = request.GET.get('ordem', 'asc')
    somente_baixo = request.GET.get('estoque_baixo') == '1'

    produtos = Produto.objects.all()
    if busca:
        produtos = produtos.filter(
            Q(descricao__icontains=busca) |
            Q(codigo_barras__icontains=busca)
        )
    if status:
        produtos = produtos.filter(status=status)
    if filtro_area:
        produtos = produtos.filter(area__nome__iexact=filtro_area)
    if ordenar_por in ['id', 'validade', 'descricao', 'criado_em']:
        exp = f"-{ordenar_por}" if ordem == 'desc' else ordenar_por
        produtos = produtos.order_by(exp)

    # pega o mínimo configurado por área
    configs = ConfiguracaoEstoque.objects.all()
    minimo_por_area = {
        cfg.area_id: cfg.estoque_minimo
        for cfg in configs if cfg.area_id is not None
    }

    hoje = now().date()

    # 0) total reservado por (codigo_barras, area_id)
    pendentes = ItemPedido.objects.filter(
        produto__codigo_barras__in=produtos.values_list('codigo_barras', flat=True),
        produto__area_id__in=produtos.values_list('area_id', flat=True),
        pedido__status__in=PENDING_STATUSES
    )
    pendentes = pendentes.filter(pedido__data_solicitacao__date__lte=hoje)
    reservas_por_combo = {
        (r['produto__codigo_barras'], r['produto__area_id']): r['total']
        for r in pendentes
                 .values('produto__codigo_barras', 'produto__area_id')
                 .annotate(total=Sum('quantidade'))
    }

    produtos_filtrados = []

    # 1) itera lote a lote, usando p.quantidade como estoque real
    for p in produtos.order_by('validade', 'criado_em'):
        # dias até vencer e mínimo
        p.dias_para_vencer = (p.validade - hoje).days if p.validade else 9999
        p.minimo_area      = minimo_por_area.get(p.area_id, 0)

        key = (p.codigo_barras, p.area_id)
        reservado_restante = reservas_por_combo.get(key, 0)

        # estoque real por lote
        real = p.quantidade

        # reserva alocada a este lote
        reservado_lote = min(real, reservado_restante)
        reservas_por_combo[key] = reservado_restante - reservado_lote

        # disponível projetado do lote
        disponivel = real - reservado_lote

        # atribuições para template
        p.estoque_real                 = real
        p.estoque_reservado            = reservado_lote
        p.estoque_disponivel_projetado = disponivel

        p.estoque_baixo = (
            disponivel <= p.minimo_area
            and p.minimo_area > 0
        )
        p.percentual_estoque = (
            (disponivel / p.minimo_area) * 100
            if p.minimo_area else 100
        )

        if not somente_baixo or p.estoque_baixo:
            produtos_filtrados.append(p)

    context = {
        'produtos': produtos_filtrados,
        'busca': busca,
        'filtro_status': status,
        'filtro_area': filtro_area,
        'ordenar_por': ordenar_por,
        'ordem': ordem,
        'total_produtos': Produto.objects.count(),
        'total_filtrados': len(produtos_filtrados),
        'areas': list(Area.objects.values('id', 'nome')),
        'fornecedores': list(Fornecedor.objects.values('id', 'nome')),
        'estoque_baixo_aplicado': somente_baixo,
    }
    return render(request, 'core/lista_produtos.html', context)



    context = {
        'produtos': produtos_filtrados,
        'busca': busca,
        'filtro_status': status,
        'filtro_area': filtro_area,
        'ordenar_por': ordenar_por,
        'ordem': ordem,
        'total_produtos': Produto.objects.count(),
        'total_filtrados': len(produtos_filtrados),
        'areas': list(Area.objects.values('id', 'nome')),
        'fornecedores': list(Fornecedor.objects.values('id', 'nome')),
        'estoque_baixo_aplicado': somente_baixo,
    }
    return render(request, 'core/lista_produtos.html', context)


    context = {
        'produtos': produtos_filtrados,
        'busca': busca,
        'filtro_status': status,
        'filtro_area': filtro_area,
        'ordenar_por': ordenar_por,
        'ordem': ordem,
        'total_produtos': Produto.objects.count(),
        'total_filtrados': len(produtos_filtrados),
        'areas': list(Area.objects.values('id', 'nome')),
        'fornecedores': list(Fornecedor.objects.values('id', 'nome')),
        'estoque_baixo_aplicado': somente_baixo,
    }
    return render(request, 'core/lista_produtos.html', context)



@login_required
def cadastro_produtos(request):
    areas = Area.objects.all()
    fornecedores = Fornecedor.objects.all()

    if request.method == "POST":
        try:
            # --- Importação em massa via confirmação de NFe ---
            codigos        = request.POST.getlist("codigo_barras")
            if codigos:
                descricoes     = request.POST.getlist("descricao")
                nomes_forneced = request.POST.getlist("fornecedor_nome")
                area_ids       = request.POST.getlist("area")
                validades      = request.POST.getlist("validade")
                quantidades    = request.POST.getlist("quantidade")
                precos         = request.POST.getlist("preco_unitario")
                statuses       = request.POST.getlist("status")
                nfe_numeros    = request.POST.getlist("nfe_numero")

                # iterar sem estourar índice
                for idx, (cb, desc, f_nome, aid, v_str, q_str, p_str, st) in enumerate(
                    zip_longest(
                        codigos, descricoes, nomes_forneced, area_ids,
                        validades, quantidades, precos, statuses,
                        fillvalue=None
                    )
                ):
                    validade = parse_date(v_str) if v_str else None
                    quantidade = int(q_str) if q_str else 0
                    preco_unitario = float(p_str) if p_str else 0.0
                    status = st or "ativo"
                    nfe_numero = nfe_numeros[idx] if idx < len(nfe_numeros) else ""

                    # busca ou cria fornecedor
                    if f_nome:
                        fornecedor, _ = Fornecedor.objects.get_or_create(nome=f_nome)
                    else:
                        fornecedor = None

                    produto = Produto(
                        nfe_numero      = nfe_numero,
                        codigo_barras   = cb or "",
                        descricao       = desc or "",
                        fornecedor      = fornecedor,
                        area_id         = aid or None,
                        validade        = validade,
                        quantidade      = quantidade,
                        preco_unitario  = preco_unitario,
                        status          = status,
                        criado_por      = request.user,
                    )
                    produto.save()

                messages.success(request, "Produtos importados com sucesso!")
                return redirect("lista_produtos")

            # --- Cadastro individual (fallback) ---
            codigo_barras = request.POST.get("codigo_barras")
            if codigo_barras:
                descricao      = request.POST.get("descricao", "")
                fornecedor_id  = request.POST.get("fornecedor")
                area_id        = request.POST.get("area")
                validade       = parse_date(request.POST.get("validade")) or None
                quantidade     = int(request.POST.get("quantidade", 0))
                preco_unitario = float(request.POST.get("preco_unitario", 0))
                status         = request.POST.get("status") or "ativo"

                if not fornecedor_id or not area_id:
                    messages.error(request, "Fornecedor e área são obrigatórios.")
                    return redirect("cadastro_produtos")

                produto = Produto(
                    codigo_barras   = codigo_barras,
                    descricao       = descricao,
                    fornecedor_id   = fornecedor_id,
                    area_id         = area_id,
                    validade        = validade,
                    quantidade      = quantidade,
                    preco_unitario  = preco_unitario,
                    status          = status,
                    criado_por      = request.user,
                )
                produto.save()

                messages.success(request, "Produto cadastrado com sucesso!")
                return redirect("lista_produtos")

        except IntegrityError:
            messages.error(request, "Erro ao cadastrar produto. Verifique os dados.")
            return redirect("cadastro_produtos")

    # GET: exibe formulário de cadastro/importação
    return render(request, "core/cadastro_produtos.html", {
        "fornecedores": fornecedores,
        "areas": areas,
    })

@csrf_exempt
@login_required
def salvar_produto_inline(request):
    if request.method == 'POST':
        try:
            dados = json.loads(request.body)
            fornecedor = Fornecedor.objects.get(id=dados['fornecedor'])
            area = Area.objects.get(id=dados['area']) if dados.get('area') else None

            produtos_iguais = Produto.objects.filter(codigo_barras=dados['codigo_barras'])
            if produtos_iguais.exists():
                maior_lote = max([int(p.lote) for p in produtos_iguais if str(p.lote).isdigit()] or [0])
                lote = maior_lote + 1
            else:
                lote = 1

            produto = Produto.objects.create(
                codigo_barras=dados['codigo_barras'],
                descricao=dados['descricao'],
                fornecedor=fornecedor,
                area=area,
                lote=str(lote),
                validade=dados['validade'],
                quantidade=int(dados['quantidade']),
                preco_unitario=float(dados['preco_unitario']),
                criado_por=request.user
            )

            return JsonResponse({
                'sucesso': True,
                'id': produto.id,
                'fornecedor_nome': fornecedor.nome,
                'area_nome': area.nome if area else '',
            })

        except Exception as e:
            return JsonResponse({'sucesso': False, 'erro': str(e)})

    return JsonResponse({'sucesso': False, 'erro': 'Método inválido'})


def editar_produto(request, produto_id):
    produto = get_object_or_404(Produto, id=produto_id)
    form = ProdutoForm(request.POST or None, instance=produto)
    if form.is_valid():
        form.save()
        return redirect('lista_produtos')
    return render(request, 'core/editar_produto.html', {
        'form': form, 'produto': produto
    })


def is_admin(user):
    return user.is_authenticated and user.papel == 'admin'

@login_required
@user_passes_test(is_admin)
def deletar_produto(request, produto_id):
    produto = get_object_or_404(Produto, id=produto_id)
    produto.delete()
    messages.success(request, "Produto excluído com sucesso.")
    return redirect('lista_produtos')

# ---------- Cosmos API ----------

def buscar_nome_produto_por_codigo(codigo_barras):
    url = f"https://api.cosmos.bluesoft.com.br/gtins/{codigo_barras}.json"
    headers = {
        "X-Cosmos-Token": "3fTpL-M47SqFLJ8qq1RAPg",  # Substitua por seu token real
        "Content-Type": "application/json",
        "User-Agent": "Cosmos-API-Request"
    }
    try:
        response = requests.get(url, headers=headers, timeout=5)
        if response.status_code == 200:
            data = response.json()
            return data.get("description", "Descrição não encontrada")
        elif response.status_code == 404:
            return "Produto não encontrado"
        else:
            return f"Erro: {response.status_code}"
    except Exception as e:
        print(f"[ERRO Cosmos API] {e}")
        return "Erro na consulta"

@login_required
def buscar_nome_produto_view(request, codigo):
    nome = buscar_nome_produto_por_codigo(codigo)
    return JsonResponse({"codigo": codigo, "nome_produto": nome})

@csrf_exempt
@login_required
def registrar_movimentacao(request):
    if request.method == 'POST':
        codigo = request.POST.get('codigo_barras')
        tipo = request.POST.get('tipo')
        quantidade = int(request.POST.get('quantidade'))
        try:
            produto = Produto.objects.get(codigo_barras=codigo)
        except Produto.DoesNotExist:
            return JsonResponse({'erro': 'Produto não encontrado'}, status=404)
        if tipo == "saida" and produto.quantidade < quantidade:
            return JsonResponse({'erro': 'Estoque insuficiente'}, status=400)
        produto.quantidade = produto.quantidade + quantidade if tipo == "entrada" else produto.quantidade - quantidade
        produto.save()
        MovimentacaoEstoque.objects.create(tipo=tipo, usuario=request.user, quantidade=quantidade, produto=produto)
        return JsonResponse({'mensagem': 'Movimentação registrada com sucesso'})
    return JsonResponse({'erro': 'Método inválido'}, status=405)

@login_required
def exportar_dashboard_excel(request):
    """
    Exporta TODAS as movimentações de estoque como CSV.
    """
    # pega todas as movimentações, em ordem cronológica decrescente
    movimentacoes = (
        MovimentacaoEstoque.objects
        .select_related('produto', 'usuario')
        .order_by('-data')
    )

    # prepara resposta CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="movimentacoes_estoque.csv"'

    writer = csv.writer(response)
    # cabeçalho
    writer.writerow(['Tipo', 'Produto', 'Quantidade', 'Usuário', 'Data/Hora'])

    # escreve cada movimentação
    for m in movimentacoes:
        writer.writerow([
            m.get_tipo_display(),
            m.produto.descricao,
            m.quantidade,
            m.usuario.username,
            m.data.strftime('%d/%m/%Y %H:%M'),
        ])

    return response

@login_required
def fornecedores_view(request):
    if request.method == "POST":
        Fornecedor.objects.create(
            nome=request.POST["nome"],
            cnpj=request.POST["cnpj"],
            endereco=request.POST["endereco"],
            telefone=request.POST["telefone"],
            email=request.POST["email"]
        )
    fornecedores = Fornecedor.objects.all()
    return render(request, "core/fornecedores.html", {"fornecedores": fornecedores})

@login_required
def ativar_fornecedor(request, pk):
    f = get_object_or_404(Fornecedor, pk=pk)
    f.ativo = True
    f.save()
    messages.success(request, f"Fornecedor {f.nome} ativado com sucesso.")
    return redirect('fornecedores')


@login_required
def desativar_fornecedor(request, pk):
    f = get_object_or_404(Fornecedor, pk=pk)
    f.ativo = False
    f.save()
    messages.success(request, f"Fornecedor {f.nome} desativado com sucesso.")
    return redirect('fornecedores')


@login_required
def deletar_fornecedor(request, pk):
    f = get_object_or_404(Fornecedor, pk=pk)

    # Verifica vínculos com produtos ativos
    if Produto.objects.filter(fornecedor=f, status='ativo').exists():
        messages.error(
            request,
            "Não foi possível excluir: existem produtos ativos vinculados a este fornecedor."
        )
        return redirect('fornecedores')

    f.delete()
    messages.success(request, f"Fornecedor {f.nome} excluído com sucesso.")
    return redirect('fornecedores')

@csrf_exempt
@require_POST
@login_required
def salvar_fornecedor_inline(request):
    try:
        data = json.loads(request.body)
        fornecedor = Fornecedor.objects.create(
            nome     = data.get("nome", ""),
            cnpj     = data.get("cnpj", ""),
            telefone = data.get("telefone", ""),
            email    = data.get("email", ""),
            ativo    = True
        )
        return JsonResponse({"sucesso": True, "id": fornecedor.id})
    except Exception as e:
        return JsonResponse({"sucesso": False, "erro": str(e)})
    
@login_required
def editar_fornecedor(request, pk):
    f = get_object_or_404(Fornecedor, pk=pk)
    if request.method == "POST":
        f.nome      = request.POST["nome"]
        f.cnpj      = request.POST["cnpj"]
        f.endereco  = request.POST.get("endereco", f.endereco)
        f.telefone  = request.POST.get("telefone", f.telefone)
        f.email     = request.POST.get("email", f.email)
        f.save()
        messages.success(request, f"Fornecedor {f.nome} atualizado com sucesso.")
        return redirect('fornecedores')
    return render(request, "core/editar_fornecedor.html", {"fornecedor": f})


# ---------- Usuários ----------

@login_required
def lista_usuarios(request):
    usuarios = User.objects.all()
    areas = Area.objects.all()
    return render(request, 'core/usuarios.html', {
        'usuarios': usuarios,
        'areas': areas,
        'total_areas': areas.count(),  
    })

@login_required
def ativar_usuario(request, usuario_id):
    if not request.user.is_superuser:
        messages.error(request, "Somente administradores podem ativar usuários.")
        return redirect('lista_usuarios')
    u = get_object_or_404(User, id=usuario_id)
    u.ativo = True
    u.save()
    messages.success(request, f"Usuário {u.username} ativado com sucesso.")
    return redirect('lista_usuarios')

@login_required
def desativar_usuario(request, usuario_id):
    if not request.user.is_superuser:
        messages.error(request, "Somente administradores podem desativar usuários.")
        return redirect('lista_usuarios')

    u = get_object_or_404(User, id=usuario_id)

    if u.papel.lower() == 'admin':
        total_admins_ativos = User.objects.filter(papel__iexact='admin', ativo=True).count()
        if total_admins_ativos <= 1:
            messages.error(request, "Não é possível desativar o único administrador ativo do sistema.")
            return redirect('lista_usuarios')

    u.ativo = False
    u.save()
    messages.success(request, f"Usuário {u.username} desativado com sucesso.")
    return redirect('lista_usuarios')


@login_required
def editar_usuario(request, usuario_id):
    u = get_object_or_404(User, id=usuario_id)
    if request.method == "POST":
        u.username  = request.POST["username"]
        u.email     = request.POST["email"]
        u.matricula = request.POST["matricula"]
        u.papel     = request.POST["papel"]
        # recebe lista de áreas
        area_ids = request.POST.getlist("areas")
        u.save()
        u.areas.set(area_ids)          # atualiza M2M
        messages.success(request, f"Usuário {u.username} atualizado com sucesso.")
        return redirect('lista_usuarios')

    return render(request, 'core/editar_usuario.html', {
        'usuario': u,
        'areas':   Area.objects.all()
    })

@login_required
def deletar_usuario(request, usuario_id):
    if not request.user.is_superuser:
        messages.error(request, "Somente administradores podem excluir usuários.")
        return redirect('lista_usuarios')
    
    u = get_object_or_404(User, id=usuario_id)

    if u.papel.lower() == 'admin':
        total_admins = User.objects.filter(papel__iexact='admin').count()
        if total_admins <= 1:
            messages.error(request, "Não é possível excluir o único administrador do sistema.")
            return redirect('lista_usuarios')

    u.delete()
    messages.success(request, f"Usuário {u.username} excluído com sucesso.")
    return redirect('lista_usuarios')


@csrf_exempt
@require_POST
@login_required
def salvar_usuario_inline(request):
    try:
        data = json.loads(request.body)
        
        # Verifica se o usuário já existe
        if User.objects.filter(username=data['username']).exists():
            return JsonResponse({'sucesso': False, 'erro': 'Usuário já existe'})

        # Cria o usuário com senha padrão
        u = User.objects.create_user(
            username=data['username'],
            email=data.get('email', ''),
            password='senha123'
        )

        # Atribui matrícula
        u.matricula = data.get('matricula', '')

        # Mapeia papéis aceitos com variações
        papel_recebido = data.get('papel', '').strip().lower()
        mapa_papeis = {
            'admin': 'admin',
            'administrador': 'admin',
            'operador': 'operador',
            'técnico': 'tecnico',  # com acento
            'tecnico': 'tecnico',  # sem acento
        }

        papel_final = mapa_papeis.get(papel_recebido)
        if not papel_final:
            return JsonResponse({'sucesso': False, 'erro': 'Papel inválido ou não selecionado'})

        u.papel = papel_final

        # Associa áreas
        area_ids = data.get('areas', [])
        if 'all' in area_ids or u.papel in ['admin', 'tecnico']:
            u.areas.set(Area.objects.values_list('id', flat=True))
        else:
            u.areas.set(area_ids)

        u.save()
        return JsonResponse({'sucesso': True})
    
    except Exception as e:
        return JsonResponse({'sucesso': False, 'erro': str(e)})


@login_required
def editar_perfil(request):
    if request.method == 'POST':
        form = ProfileForm(request.POST, instance=request.user)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, "Perfil atualizado com sucesso.")
            return redirect('editar_perfil')
        else:
            messages.error(request, "Por favor, corrija os erros abaixo.")
    else:
        form = ProfileForm(instance=request.user)

    return render(request, 'core/editar_perfil.html', {
        'form': form
    })

@login_required
def dashboard(request):
    # 1) Pega o parâmetro 'data' (YYYY‑MM‑DD). Se inválido ou ausente, usa hoje.
    data_str = request.GET.get('data')
    if data_str:
        try:
            data_mov = datetime.strptime(data_str, '%Y-%m-%d').date()
        except ValueError:
            data_mov = date.today()
    else:
        data_mov = date.today()

    # 2) Contagens básicas
    produtos_count = Produto.objects.count()
    pedidos_count  = Pedido.objects.count()
    usuarios_count = Usuario.objects.count()

    # 3) Pedidos por status
    pedidos_por_status = {
        status: Pedido.objects.filter(status=status).count()
        for status, _ in Pedido.STATUS_CHOICES
    }

    # 4) Entradas totais (quantidade inicial cadastrada)
    entradas_totais = Produto.objects.aggregate(
        total=Sum('quantidade_inicial')
    )['total'] or 0

    # 5) Em Estoque (estoque real): soma de estoque_info.real por combo código+área
    estoque_real_total = 0
    combos = Produto.objects.values('codigo_barras', 'area_id').distinct()
    for combo in combos:
        p = Produto.objects.filter(
            codigo_barras=combo['codigo_barras'],
            area_id=combo['area_id']
        ).first()
        if p:
            estoque_real_total += p.estoque_info(data_limite=data_mov)['real']

    # 6) Valor total em estoque (quantidade * preço_unitário)
    valor_total = Produto.objects.aggregate(
        valor=Sum(
            F('quantidade') * F('preco_unitario'),
            output_field=DecimalField(max_digits=20, decimal_places=2)
        )
    )['valor'] or 0

    # 7) Movimentações e logs do dia
    movimentacoes_dia = (
        MovimentacaoEstoque.objects
        .select_related('produto', 'usuario')
        .filter(data__date=data_mov)
        .order_by('-data')
    )
    logs_dia = LogAcao.objects.filter(data_hora__date=data_mov).order_by('-data_hora')

    return render(request, 'core/dashboard.html', {
        'produtos_count':       produtos_count,
        'pedidos_count':        pedidos_count,
        'usuarios_count':       usuarios_count,
        'pedidos_por_status':   pedidos_por_status,
        'entradas_totais':      entradas_totais,
        'estoque_real_total':   estoque_real_total,
        'valor_total':          valor_total,
        'movimentacoes_dia':    movimentacoes_dia,
        'logs_dia':             logs_dia,
        'data_mov':             data_mov,
        'STATUS_CHOICES':       Pedido.STATUS_CHOICES,
    })

    
# ---------- Importação NFe ----------

@login_required
def upload_nfe(request):
    if request.method == "POST" and request.FILES.get("arquivo"):
        xml_file = request.FILES["arquivo"]
        fs = FileSystemStorage()
        filepath = fs.path(fs.save(xml_file.name, xml_file))
        tree = ET.parse(filepath)
        root = tree.getroot()
        ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}

        # 1) EXTRAIR O NÚMERO DA NFe
        nfe_numero = root.findtext(".//nfe:ide/nfe:nNF", default="", namespaces=ns)

        # 2) DADOS DO EMITENTE
        emitente = root.find(".//nfe:emit", ns)
        fornecedor_nome = emitente.findtext("nfe:xNome", default="", namespaces=ns)
        fornecedor_cnpj  = emitente.findtext("nfe:CNPJ", default="", namespaces=ns)
        fornecedor, _ = Fornecedor.objects.get_or_create(
            cnpj=fornecedor_cnpj,
            defaults={"nome": fornecedor_nome, "endereco":"", "telefone":"", "email":""}
        )

        # 3) ITENS
        produtos_extraidos = []
        for det in root.findall(".//nfe:det", ns):
            prod = det.find("nfe:prod", ns)
            if prod is None:
                continue
            produtos_extraidos.append({
                "nfe_numero":      nfe_numero,
                "codigo_barras":   prod.findtext("nfe:cProd",  default="", namespaces=ns),
                "descricao":       prod.findtext("nfe:xProd",  default="", namespaces=ns),
                "quantidade":      prod.findtext("nfe:qCom",   default="0",  namespaces=ns),
                "preco_unitario":  prod.findtext("nfe:vUnCom", default="0.00", namespaces=ns),
                "lote":            "",
                "validade":        "",
                "status":          "ativo",
                "fornecedor_nome": fornecedor.nome,     # passa o nome, não o ID
                "area_id":         None,                # será selecionada no template
            })

        return render(request, "core/confirmar_importacao.html", {
            "produtos":    produtos_extraidos,
            "areas":       Area.objects.all(),
            "nfe_numero":  nfe_numero,
        })

    return render(request, "core/upload_nfe.html")


@login_required
def exportar_produtos_excel(request):
    busca = request.GET.get("busca")
    status = request.GET.get("filtro_status")
    produtos = Produto.objects.all()

    if busca:
        produtos = produtos.filter(
            Q(descricao__icontains=busca) |
            Q(codigo_barras__icontains=busca)
        )
    if status:
        produtos = produtos.filter(status=status)

    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"

    # 1) Carimbo de data/hora na primeira linha
    timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    cell_stamp = ws.cell(row=1, column=1, value=f"Exportado em: {timestamp}")
    cell_stamp.font = Font(bold=True)
    cell_stamp.alignment = Alignment(horizontal="right")

    # 2) Cabeçalho (linha 2)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "Código de Barras", "Lote", "Descrição", "Fornecedor", "Área",
        "Validade", "Quantidade", "Estoque Real", "Reservado", "Disponível",
        "Preço Unitário", "Usuário", "Data de Cadastro"
    ]
    for col_num, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # 3) Dados dos produtos (a partir da linha 3)
    for idx, produto in enumerate(produtos, start=3):
        info = produto.estoque_info(data_limite=None)  # ou passe data se desejar histórico
        ws.cell(row=idx, column=1, value=produto.codigo_barras)
        ws.cell(row=idx, column=2, value=produto.lote)
        ws.cell(row=idx, column=3, value=produto.descricao)
        ws.cell(row=idx, column=4, value=getattr(produto.fornecedor, 'nome', ''))
        ws.cell(row=idx, column=5, value=produto.area.nome if produto.area else "")
        ws.cell(row=idx, column=6, value=produto.validade.strftime('%d/%m/%Y') if produto.validade else "")
        ws.cell(row=idx, column=7, value=produto.quantidade)
        ws.cell(row=idx, column=8, value=info['real'])
        ws.cell(row=idx, column=9, value=info['reservado'])
        ws.cell(row=idx, column=10, value=info['disponivel'])
        ws.cell(row=idx, column=11, value=float(produto.preco_unitario))
        ws.cell(row=idx, column=12, value=produto.criado_por.username if produto.criado_por else "")
        ws.cell(row=idx, column=13, value=produto.criado_em.strftime('%d/%m/%Y %H:%M') if produto.criado_em else "")

    # 4) Ajustar largura das colunas
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # 5) Resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=produtos.xlsx'
    wb.save(response)
    return response

@login_required
def exportar_pedidos_excel(request):
    pedidos = Pedido.objects.select_related('usuario', 'aprovado_por')\
                            .prefetch_related('itens__produto__area', 'subitens__produto__area')\
                            .order_by('-data_solicitacao')

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "Código", "Data Solicitação", "Solicitante", "Email", "Necessário em", "Status",
        "Aprovado por", "Data Aprovação", "Separado em", "Retirado por", "Data Retirada",
        "Produto", "Código de Barras", "Área", "Qtd Solicitada", "Qtd Liberada", "Estoque Disp.", "Observação"
    ]
    ws.append(headers)

    for col_num, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    row = 2
    for pedido in pedidos:
        itens = list(pedido.itens.all())
        subitens = list(pedido.subitens.all())
        total_linhas = len(itens) + len(subitens)
        start_row = row

        for item in itens:
            produto = item.produto
            ws.append([
                "", "", "", "", "", "", "", "", "", "", "",
                produto.descricao if produto else "",
                produto.codigo_barras if produto else "",
                produto.area.nome if produto and produto.area else "",
                item.quantidade,
                item.liberado if item.liberado is not None else item.quantidade,
                item.estoque_no_pedido or "",
                item.observacao or ""
            ])
            row += 1

        for subitem in subitens:
            produto = subitem.produto
            ws.append([
                "", "", "", "", "", "", "", "", "", "", "",
                f"{produto.descricao} (subitem)" if produto else "",
                produto.codigo_barras if produto else "",
                produto.area.nome if produto and produto.area else "",
                subitem.quantidade,
                "—",
                subitem.estoque_no_pedido or "",
                ""
            ])
            row += 1

        if total_linhas == 0:
            ws.append([
                pedido.codigo,
                pedido.data_solicitacao.strftime('%d/%m/%Y %H:%M') if pedido.data_solicitacao else "",
                pedido.usuario.username if pedido.usuario else "",
                pedido.usuario.email if pedido.usuario else "",
                pedido.data_necessaria.strftime('%d/%m/%Y') if pedido.data_necessaria else "",
                pedido.get_status_display(),
                pedido.aprovado_por.username if pedido.aprovado_por else "",
                pedido.data_aprovacao.strftime('%d/%m/%Y %H:%M') if pedido.data_aprovacao else "",
                pedido.data_separacao.strftime('%d/%m/%Y %H:%M') if pedido.data_separacao else "",
                pedido.retirado_por or "",
                pedido.data_retirada.strftime('%d/%m/%Y %H:%M') if pedido.data_retirada else "",
            ] + [""] * 7)
            row += 1
        else:
            for col, val in enumerate([
                pedido.codigo,
                pedido.data_solicitacao.strftime('%d/%m/%Y %H:%M') if pedido.data_solicitacao else "",
                pedido.usuario.username if pedido.usuario else "",
                pedido.usuario.email if pedido.usuario else "",
                pedido.data_necessaria.strftime('%d/%m/%Y') if pedido.data_necessaria else "",
                pedido.get_status_display(),
                pedido.aprovado_por.username if pedido.aprovado_por else "",
                pedido.data_aprovacao.strftime('%d/%m/%Y %H:%M') if pedido.data_aprovacao else "",
                pedido.data_separacao.strftime('%d/%m/%Y %H:%M') if pedido.data_separacao else "",
                pedido.retirado_por or "",
                pedido.data_retirada.strftime('%d/%m/%Y %H:%M') if pedido.data_retirada else ""
            ], start=1):
                ws.merge_cells(start_row=start_row, end_row=row-1, start_column=col, end_column=col)
                ws.cell(row=start_row, column=col, value=val).alignment = center

    # Auto ajuste largura
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=pedidos.xlsx'
    wb.save(response)
    return response

@login_required
@user_passes_test(is_admin)
def exportar_estoque_por_area_excel(request):
    """
    Exporta um CSV com cada Área e seu estoque disponível (usando estoque_disponivel até hoje).
    """
    hoje = date.today()
    # Pega todas as áreas
    areas = Area.objects.all()

    # Prepara resposta CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="estoque_por_area.csv"'

    writer = csv.writer(response)
    writer.writerow(['Área', 'Estoque Disponível'])

    # Para cada área, soma o estoque_disponivel de cada produto pertencente
    for area in areas:
        # busca um produto de referência para chamar estoque_disponivel; mas queremos somar
        total_disponivel = 0
        # pega todos os combos de código nessa área
        combos = (
            Produto.objects
            .filter(area=area)
            .values('codigo_barras')
            .distinct()
        )
        for combo in combos:
            p = (
                Produto.objects
                .filter(codigo_barras=combo['codigo_barras'], area=area)
                .first()
            )
            if p:
                total_disponivel += p.estoque_disponivel(hoje)
        writer.writerow([area.nome, total_disponivel])

    return response

@login_required
def lista_logs(request):
    if not request.user.is_superuser:
        return redirect('lista_produtos')
    logs = LogAcao.objects.all().order_by('-data_hora')
    return render(request, "core/logs.html", {"logs": logs})

@login_required
@transaction.atomic
def novo_pedido(request):
    # monta áreas e mínimos
    if request.user.papel in ['admin', 'tecnico']:
        areas = Area.objects.all()
    else:
        areas = request.user.areas.all()
    area_minimos = {a.id: a.estoque_minimo for a in areas}

    if request.method == "POST":
        # 1) Gera código
        ultimo = Pedido.objects.aggregate(Max("id"))["id__max"] or 0
        codigo = f"PD{str(ultimo + 1).zfill(5)}"

        # 2) Status e aprovação
        if request.user.papel in ["admin", "tecnico"]:
            status_inicial = "aprovado"
            aprovado_por = request.user
            data_aprovacao = now()
        else:
            status_inicial = "aguardando_aprovacao"
            aprovado_por = None
            data_aprovacao = None

        # 3) Cria o pedido
        data_necessaria_pedido = request.POST.get('data_necessaria') or None
        pedido = Pedido.objects.create(
            codigo=codigo,
            usuario=request.user,
            status=status_inicial,
            aprovado_por=aprovado_por,
            data_aprovacao=data_aprovacao,
            data_necessaria=data_necessaria_pedido,
        )

        hoje = now().date()
        area_ids    = request.POST.getlist('area_id')
        produto_ids = request.POST.getlist('produto_id')
        quantidades = request.POST.getlist('quantidade')
        observacoes = request.POST.getlist('obs_item')

        # 4) Loop sobre itens
        for area_id, prod_id, qt, obs in zip_longest(
                area_ids, produto_ids, quantidades, observacoes, fillvalue=''):
            if not (area_id and prod_id and qt):
                continue

            # valida quantidade
            try:
                quantidade = int(qt)
                if quantidade <= 0:
                    raise ValueError()
            except (ValueError, TypeError):
                messages.error(request, f"Quantidade inválida para o produto {prod_id}.")
                continue

            # bloqueio para consistência
            inst = get_object_or_404(
                Produto.objects.select_for_update(),
                pk=int(prod_id)
            )

            # confirma área
            if inst.area_id != int(area_id):
                messages.error(
                    request,
                    f"Produto {inst.descricao} não pertence à área selecionada."
                )
                continue  # removido raise para evitar 500

            # obtém estoque projetado
            info = inst.estoque_info(data_limite=hoje)
            estoque_disp = info.get('disponivel', 0)

            if estoque_disp < quantidade:
                messages.error(
                    request,
                    f"Estoque insuficiente para {inst.descricao} na área {inst.area.nome}. "
                    f"Disponível: {estoque_disp}, Solicitado: {quantidade}."
                )
                continue  # removido raise para evitar 500

            # grava o item
            ItemPedido.objects.create(
                pedido=pedido,
                produto=inst,
                quantidade=quantidade,
                observacao=obs.strip(),
                estoque_no_pedido=estoque_disp
            )

        # 5) Feedback e redirect
        if status_inicial == "aprovado":
            messages.success(request, "Pedido aprovado com sucesso!")
        else:
            messages.success(request, "Pedido registrado com sucesso! Aguardando aprovação.")

        return redirect(reverse('lista_pedidos'))

    # GET: renderiza o formulário
    return render(request, "core/novo_pedido.html", {
        "areas": areas,
        "areas_json": json.dumps(
            list(areas.values("id", "nome")),
            cls=DjangoJSONEncoder
        ),
        "area_minimos": area_minimos,
    })

@login_required
@transaction.atomic
def aprovar_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)

    if pedido.status != 'aguardando_aprovacao':
        messages.warning(request, "Este pedido já foi aprovado.")
        return redirect('detalhes_pedido', pedido_id=pedido.id)

    for item in pedido.itens.all():
        qtd_restante = item.quantidade
        produtos = Produto.objects.filter(
            codigo_barras=item.produto.codigo_barras,
            area=item.produto.area,
            quantidade__gt=0
        ).order_by('criado_em', 'validade')

        for prod in produtos:
            if qtd_restante == 0:
                break
            if prod.quantidade >= qtd_restante:
                prod.quantidade -= qtd_restante
                prod.save()
                qtd_restante = 0
            else:
                qtd_restante -= prod.quantidade
                prod.quantidade = 0
                prod.save()

        if qtd_restante > 0:
            messages.error(request, f"Estoque insuficiente para o produto {item.produto.descricao}")
            transaction.set_rollback(True)
            return redirect('detalhes_pedido', pedido_id=pedido.id)

        # Atualiza o item.produto para refletir a nova soma
        item.produto.quantidade = Produto.objects.filter(
            codigo_barras=item.produto.codigo_barras,
            area=item.produto.area
        ).aggregate(total=Sum('quantidade'))['total'] or 0
        item.produto.save(update_fields=['quantidade'])

    pedido.aprovar(request.user)
    pedido.data_aprovacao = now()
    pedido.save()

    messages.success(request, "Pedido aprovado com sucesso. Estoque atualizado.")
    return redirect('detalhes_pedido', pedido_id=pedido.id)



@login_required
def separar_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    if pedido.status == 'aprovado':
        pedido.marcar_como_separado()
        if pedido.usuario.email:
            send_mail(
                subject='Pedido separado!',
                message=f'Seu pedido {pedido.codigo} foi separado e está disponível para retirada.',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[pedido.usuario.email],
                fail_silently=True
            )
        messages.success(request, f"Pedido {pedido.codigo} marcado como separado.")
    else:
        messages.error(request, "Pedido não está no status correto para separação.")
    return redirect('lista_pedidos')

@login_required
def registrar_retirada(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    if request.method == 'POST':
        nome = request.POST.get('retirado_por')
        if nome:
            pedido.registrar_retirada(nome)
            messages.success(request, f"Pedido {pedido.codigo} registrado como retirado.")
        else:
            messages.error(request, "Informe quem retirou o pedido.")
    return redirect('lista_pedidos')

def is_admin_tecnico(u):
    return u.papel in ['admin', 'tecnico']

@login_required
def lista_pedidos(request):
    status = request.GET.get("status")
    ordenar_por = request.GET.get("ordenar_por", "-data_solicitacao")  # padrão decrescente
    pagina = request.GET.get("page")

    # Filtra pedidos por papel do usuário
    if request.user.papel in ['admin', 'tecnico']:
        pedidos_qs = Pedido.objects.all()
    else:
        pedidos_qs = Pedido.objects.filter(usuario=request.user)

    # Filtro por status
    if status:
        pedidos_qs = pedidos_qs.filter(status=status)

    # Ordenação (campo pode ser passado com "-" para ordem decrescente)
    pedidos_qs = pedidos_qs.order_by(ordenar_por)

    # Paginação
    paginator = Paginator(pedidos_qs, 10)
    pedidos = paginator.get_page(pagina)

    return render(request, 'core/lista_pedidos.html', {
        'pedidos': pedidos,
        'STATUS_CHOICES': Pedido.STATUS_CHOICES,
        'areas': Area.objects.all(),
        'produtos': Produto.objects.filter(status='ativo'),
        'status_selecionado': status,
        'ordenar_por': ordenar_por.lstrip('-'),
        'ordem': 'desc' if ordenar_por.startswith('-') else 'asc',
    })


@login_required
@transaction.atomic
def detalhe_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    is_admin = request.user.papel == 'admin'

    # Bloqueia POST para não-admins
    if request.method == 'POST' and not is_admin:
        return redirect('lista_pedidos')

    if request.method == 'POST':
        action = request.POST.get('action')

        # Aprovar pedido → dispara a baixa de estoque em todos os lotes
        if action == 'approve' and pedido.status == 'aguardando_aprovacao':
            pedido.aprovar(request.user)
            messages.success(request, "Pedido aprovado com sucesso! Estoque atualizado.")

        # Reprovar pedido
        elif action == 'reject' and pedido.status == 'aguardando_aprovacao':
            pedido.status = 'rejeitado'
            pedido.save()
            messages.success(request, "Pedido rejeitado.")

        # Separar pedido → só marca a quantidade liberada e muda status
        elif action == 'separar' and pedido.status == 'aprovado':
            for item in pedido.itens.all():
                raw = request.POST.get(f'liberado_{item.id}')
                liberado = int(raw) if raw and raw.isdigit() else 0

                # Limita ao saldo que havia no momento do pedido
                estoque_no_pedido = item.estoque_no_pedido or 0
                item.liberado = min(liberado, item.quantidade, estoque_no_pedido)
                item.save()

            pedido.marcar_como_separado()
            messages.success(request, "Pedido marcado como separado.")

        # Retirar pedido → dispara apenas o log de saída e muda status para entregue
        elif action == 'retirar' and pedido.status == 'separado':
            quem = request.POST.get('retirado_por') or request.user.username
            pedido.registrar_retirada(quem)
            messages.success(request, "Pedido registrado como retirado.")

        else:
            messages.error(request, "Ação não permitida ou estado inválido.")

        return redirect('lista_pedidos')

    # GET: exibe lista de pedidos
    pedidos = Pedido.objects.select_related('usuario').prefetch_related('itens__produto')
    return render(request, 'core/lista_pedidos.html', {
        'pedidos': pedidos,
        'status_selecionado': request.GET.get('status', ''),
        'is_admin': is_admin,
        'STATUS_CHOICES': Pedido.STATUS_CHOICES,
    })
    
def detalhes_pedido_view(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    return render(request, 'core/detalhes_pedido.html', {'pedido': pedido})

def deletar_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    pedido.delete()
    messages.success(request, f"Pedido {pedido.codigo} excluído com sucesso.")
    return redirect('lista_pedidos')


@login_required
@user_passes_test(is_admin)
def configuracoes_view(request):
    # --- Dados principais ---
    areas = list(Area.objects.all())
    configuracoes = ConfiguracaoEstoque.objects.select_related('area')

    # Mapeia cada área ao seu estoque mínimo; padrão 50%
    minimo_por_area = {
        cfg.area_id: cfg.estoque_minimo
        for cfg in configuracoes if cfg.area_id
    }
    for area in areas:
        area.minimo = minimo_por_area.get(area.id, 50)

    # --- Forms iniciais ---
    form_area = AreaForm()
    form_config = ConfiguracaoEstoqueForm(initial={'estoque_minimo': 50})

    # --- Tratamento de POST ---
    if request.method == 'POST':
        # Adicionar nova área
        if 'nova_area' in request.POST:
            form_area = AreaForm(request.POST)
            if form_area.is_valid():
                form_area.save()
                messages.success(request, "Área adicionada com sucesso.")
                return redirect('configuracoes')
        # Salvar nova configuração de estoque mínimo
        elif 'nova_configuracao' in request.POST:
            form_config = ConfiguracaoEstoqueForm(request.POST)
            if form_config.is_valid():
                form_config.save()
                messages.success(request, "Configuração salva com sucesso.")
                return redirect('configuracoes')

    # --- Renderiza a página de configurações ---
    return render(request, 'core/configuracoes.html', {
        'areas': areas,
        'configuracoes': configuracoes,
        'form_area': form_area,
        'form_config': form_config,
    })

# ---------- Áreas ----------

def lista_areas(request):
    if request.method == "POST":
        nome = request.POST.get("nome")
        Area.objects.create(nome=nome)
        return redirect("lista_areas")
    return render(request, "core/areas.html", {
        "areas": Area.objects.all()
    })

@login_required
@require_GET
def produtos_por_area(request, area_id):
    qs = Produto.objects.filter(area_id=area_id, status="ativo")

    grouped = (
        qs.values("codigo_barras", "area_id")
          .annotate(
             produto_id    = Max("id"),
             descricao     = Max("descricao"),
             validade      = Min("validade"),
             total_estoque = Sum("quantidade")
          )
    )

    data = []
    for item in grouped:
        disponivel = calcular_estoque_disponivel(item["codigo_barras"], item["area_id"])
        if disponivel <= 0:
            continue

        data.append({
            "id":         item["produto_id"],
            "descricao":  item["descricao"],
            "validade":   item["validade"].strftime("%Y-%m-%d") if item["validade"] else "",
            "disponivel": disponivel,
        })

    return JsonResponse(data, safe=False)


@login_required
def exportar_log_excel(request):
    # opcional: controle de permissão
    if not request.user.is_superuser:
        return redirect('lista_produtos')

    logs = LogAcao.objects.select_related('usuario').order_by('-data_hora')
    wb = Workbook()
    ws = wb.active
    ws.title = "Logs de Auditoria"
    ws.append(["Usuário", "Ação", "Detalhes", "Data/Hora", "IP"])

    for log in logs:
        ws.append([
            log.usuario.username if log.usuario else "—",
            log.acao,
            log.detalhes or "—",
            log.data_hora.strftime("%d/%m/%Y %H:%M:%S"),
            log.ip or "—",
        ])

    filename = f"auditoria_logs_{datetime.date.today().isoformat()}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
def exportar_logs_excel(request):
    # só superuser ou admin pode baixar
    if not request.user.is_superuser:
        return redirect('lista_produtos')

    logs = LogAcao.objects.select_related('usuario').order_by('-data_hora')

    wb = Workbook()
    ws = wb.active
    ws.title = "Logs de Auditoria"

    # cabeçalho
    ws.append(["Usuário", "Ação", "Detalhes", "Data/Hora", "IP"])

    # linhas
    for log in logs:
        ws.append([
            log.usuario.username if log.usuario else "—",
            log.acao,
            log.detalhes or "—",
            log.data_hora.strftime("%d/%m/%Y %H:%M:%S"),
            log.ip or "—",
        ])

    # gera nome de arquivo com a data de hoje
    filename = f"auditoria_logs_{date.today().isoformat()}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@csrf_exempt
@login_required
def bulk_delete_produtos(request):
    """
    Recebe via POST um JSON { "ids": [1,2,3,...] }
    e deleta todos os produtos cujo id esteja nessa lista.
    """
    if request.method == 'POST':
        try:
            dados = json.loads(request.body)
            ids = dados.get('ids', [])
            if not isinstance(ids, list):
                raise ValueError("ids deve ser uma lista")
            Produto.objects.filter(id__in=ids).delete()
            return JsonResponse({'sucesso': True})
        except Exception as e:
            return JsonResponse({'sucesso': False, 'erro': str(e)}, status=400)
    return JsonResponse({'sucesso': False, 'erro': 'Método não permitido'}, status=405)

@register.filter
def saldo_para(produto, data_limite=None):
    """
    Retorna o estoque disponível do produto até a data‑limite
    (usado em lista_pedidos.html).
    """
    from datetime import datetime

    # converte string no formato 'yyyy-mm-dd' para date
    if isinstance(data_limite, str):
        try:
            data_limite = datetime.strptime(data_limite, '%Y-%m-%d').date()
        except ValueError:
            data_limite = None

    return produto.estoque_disponivel(data_limite)

def is_admin(user):
    return user.is_authenticated and user.papel == 'admin'

@login_required
@user_passes_test(is_admin)
def editar_area(request, pk):
    area = get_object_or_404(Area, pk=pk)
    if request.method == "POST":
        form = AreaForm(request.POST, instance=area)
        if form.is_valid():
            form.save()
            messages.success(request, "Área atualizada com sucesso.")
            return redirect("configuracoes")
    else:
        form = AreaForm(instance=area)
    return render(request, "core/editar_area.html", {"form": form})

@login_required
@user_passes_test(is_admin)
def deletar_area(request, pk):
    area = get_object_or_404(Area, pk=pk)
    area.delete()
    messages.success(request, "Área removida.")
    return redirect("configuracoes")

@login_required
@user_passes_test(is_admin)
def editar_configuracao(request, area_id):
    # get_or_create para não falhar quando ainda não existir
    cfg, _ = ConfiguracaoEstoque.objects.get_or_create(
        area_id=area_id,
        defaults={'estoque_minimo': 50}
    )
    if request.method == 'POST':
        form = ConfiguracaoEstoqueForm(request.POST, instance=cfg)
        if form.is_valid():
            form.save()
    return redirect('configuracoes')

@login_required
@user_passes_test(is_admin)
def deletar_configuracao(request, pk):
    cfg = get_object_or_404(ConfiguracaoEstoque, pk=pk)
    cfg.delete()
    messages.success(request, "Configuração removida.")
    return redirect("configuracoes")

def is_admin_or_tecnico(user):
    return user.is_authenticated and user.papel in ('admin', 'tecnico')

@login_required
@user_passes_test(is_admin_or_tecnico)
def lista_sessoes(request):
    sessoes = SessionLog.objects.all()
    return render(request, 'core/lista_sessoes.html', {'sessoes': sessoes})

@login_required
@user_passes_test(lambda u: u.is_superuser)
def exportar_sessoes_excel(request):
    # busca todas as sessões
    sessoes = SessionLog.objects.select_related('user').order_by('-login_time')

    wb = Workbook()
    ws = wb.active
    ws.title = "Sessões de Usuário"

    # cabeçalho
    ws.append(["Usuário", "Login", "Logout", "Duração", "IP"])

    # linhas de dados
    for s in sessoes:
        ws.append([
            s.user.username,
            s.login_time.strftime("%d/%m/%Y %H:%M:%S"),
            s.logout_time.strftime("%d/%m/%Y %H:%M:%S") if s.logout_time else "—",
            str(s.duration) if s.duration else "—",
            s.ip or "—",
        ])

    # monta resposta
    filename = f"sessoes_{date.today().isoformat()}.xlsx"
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp